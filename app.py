import os
import io
import json

from flask import Flask, request, jsonify, render_template_string
import openpyxl
from google import genai
from google.genai import types

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'}
EXCEL_FILE = 'CONSUMABLE_AUTO_CREATE_PR.xlsx'

# ─── Gemini API ───────────────────────────────────────────────────────────────
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
_gemini_client = genai.Client(api_key=GEMINI_API_KEY)

# ─── Kategori Spare Part ──────────────────────────────────────────────────────
CATEGORIES = {
    'BEARINGS':    'Bearing / bantalan mekanis',
    'LUBRICATION': 'Pelumas: Oli & Grease',
    'PENUNJANG':   'Bahan penunjang (amplas, lem, kawat las, sikat)',
    'AIRMATIC':    'Komponen pneumatik (fitting, coupler selang angin)',
    'ELECTRICAL':  'Komponen listrik (MCB, kabel, lampu, fuse)',
}

# ─── Load Excel Data ──────────────────────────────────────────────────────────
def load_spare_parts():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    data = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        items = []
        header_found = False
        for row in ws.iter_rows(values_only=True):
            if row[0] == 'NO' and row[1] == 'CODE MATERIAL':
                header_found = True
                continue
            if not header_found:
                continue
            if not row[0] or not row[1]:
                continue
            items.append({
                'no': row[0],
                'code': row[1],
                'name': row[2],
                'min_stock': row[3],
                'max_stock': row[4],
                'satuan': row[5],
            })
        data[sheet_name] = items
    return data

SPARE_PARTS = load_spare_parts()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ─── Gemini Recognition ───────────────────────────────────────────────────────
def recognize_image_gemini(image_bytes: bytes) -> dict:
    cat_list = "\n".join(f"- {cat}: {desc}" for cat, desc in CATEGORIES.items())
    prompt = f"""Kamu adalah sistem klasifikasi spare part industri.
Lihat gambar ini, identifikasi nama barangnya secara spesifik, lalu tentukan kategorinya.

Kategori yang tersedia:
{cat_list}

Jawab dalam format JSON berikut (tanpa markdown, tanpa kode block):
{{
  "item_name": "<nama barang spesifik, contoh: Bearing, Inverter, Grease, MCB, Amplas, Selang Angin>",
  "category": "<nama kategori>",
  "confidence": "<HIGH|MEDIUM|LOW>",
  "score": <angka 0-100>,
  "reason": "<alasan singkat dalam bahasa Indonesia>"
}}

Pilih SATU kategori yang paling cocok. item_name harus nama barang yang terlihat di gambar, bukan nama kategori."""

    image_part = types.Part.from_bytes(data=image_bytes, mime_type="image/jpeg")
    response = _gemini_client.models.generate_content(
        model="gemini-2.5-flash-lite",
        contents=[prompt, image_part],
    )

    raw = response.text.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    result = json.loads(raw.strip())

    category = result.get("category", "").upper()
    if category not in CATEGORIES:
        for cat in CATEGORIES:
            if cat in result.get("category", "").upper():
                category = cat
                break
        else:
            category = list(CATEGORIES.keys())[0]

    score = float(result.get("score", 50))
    scores = {cat: round((100 - score) / (len(CATEGORIES) - 1), 1) for cat in CATEGORIES}
    scores[category] = round(score, 1)

    return {
        'item_name': result.get("item_name", category),
        'category': category,
        'confidence': result.get("confidence", "MEDIUM"),
        'score': round(score, 1),
        'scores_per_category': scores,
        'reason': result.get("reason", ""),
        'item_description': CATEGORIES[category],
    }

def find_matching_items(category: str, max_items: int = 10) -> list:
    return SPARE_PARTS.get(category, [])[:max_items]

# ─── Routes ───────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template_string(HTML_UI)

@app.route('/api/recognize', methods=['POST'])
def recognize():
    if 'image' not in request.files:
        return jsonify({'error': 'Tidak ada file image'}), 400
    file = request.files['image']
    if file.filename == '':
        return jsonify({'error': 'Tidak ada file dipilih'}), 400
    if not allowed_file(file.filename):
        return jsonify({'error': f'Format tidak didukung. Gunakan: {", ".join(ALLOWED_EXTENSIONS)}'}), 400

    image_bytes = file.read()
    try:
        result = recognize_image_gemini(image_bytes)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    category = result['category']
    return jsonify({
        'recognition': result,
        'category': category,
        'category_description': CATEGORIES[category],
        'total_items_in_category': len(SPARE_PARTS.get(category, [])),
        'matched_items': find_matching_items(category),
    })

@app.route('/api/items', methods=['GET'])
def get_all_items():
    summary = {cat: len(items) for cat, items in SPARE_PARTS.items()}
    return jsonify({
        'categories': list(SPARE_PARTS.keys()),
        'total_per_category': summary,
        'grand_total': sum(summary.values()),
    })

@app.route('/api/items/<category>', methods=['GET'])
def get_items_by_category(category):
    cat = category.upper()
    if cat not in SPARE_PARTS:
        return jsonify({'error': f'Kategori "{cat}" tidak ditemukan',
                        'available': list(SPARE_PARTS.keys())}), 404
    return jsonify({'category': cat, 'total': len(SPARE_PARTS[cat]), 'items': SPARE_PARTS[cat]})

@app.route('/api/search', methods=['GET'])
def search_items():
    query = request.args.get('q', '').lower()
    category_filter = request.args.get('category', '').upper()
    results = []
    cats = [category_filter] if category_filter in SPARE_PARTS else list(SPARE_PARTS.keys())
    for cat in cats:
        for item in SPARE_PARTS[cat]:
            name = str(item.get('name', '')).lower()
            code = str(item.get('code', '')).lower()
            if query in name or query in code:
                results.append({'category': cat, **item})
    return jsonify({'query': query, 'total': len(results), 'results': results})

# ─── HTML UI ──────────────────────────────────────────────────────────────────
HTML_UI = """<!DOCTYPE html>
<html lang="id">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Spare Part Image Recognition</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'Segoe UI', sans-serif; background: #0f172a; color: #e2e8f0; min-height: 100vh; }
  .header { background: linear-gradient(135deg, #1e3a5f, #2d6a9f); padding: 24px 32px; border-bottom: 1px solid #334155; }
  .header h1 { font-size: 1.6rem; font-weight: 700; color: #60a5fa; }
  .header p { color: #94a3b8; margin-top: 4px; font-size: 0.9rem; }
  .container { max-width: 960px; margin: 0 auto; padding: 32px 16px; }
  .card { background: #1e293b; border: 1px solid #334155; border-radius: 12px; padding: 24px; margin-bottom: 24px; }
  .card h2 { font-size: 0.85rem; font-weight: 600; color: #93c5fd; margin-bottom: 16px; text-transform: uppercase; letter-spacing: 0.05em; }
  .drop-zone { border: 2px dashed #475569; border-radius: 10px; padding: 48px 24px; text-align: center; cursor: pointer; transition: all 0.2s; }
  .drop-zone:hover, .drop-zone.drag-over { border-color: #3b82f6; background: rgba(59,130,246,0.08); }
  .drop-zone input[type=file] { display: none; }
  .drop-zone .icon { font-size: 2.5rem; margin-bottom: 12px; }
  .drop-zone p { color: #94a3b8; margin-bottom: 8px; }
  .drop-zone .hint { font-size: 0.8rem; color: #64748b; }
  .btn { display: inline-flex; align-items: center; gap: 8px; padding: 10px 20px; border-radius: 8px; border: none; cursor: pointer; font-size: 0.9rem; font-weight: 600; transition: all 0.2s; }
  .btn-primary { background: #3b82f6; color: #fff; }
  .btn-primary:hover { background: #2563eb; }
  .btn-primary:disabled { background: #475569; cursor: not-allowed; }
  .preview-wrap { display: none; text-align: center; margin-bottom: 16px; }
  .preview-wrap img { max-height: 220px; border-radius: 8px; border: 1px solid #334155; }
  .result-box { display: none; }
  .badge { display: inline-block; padding: 4px 12px; border-radius: 20px; font-size: 0.75rem; font-weight: 700; letter-spacing: 0.05em; }
  .badge-BEARINGS   { background: #1e3a5f; color: #60a5fa; }
  .badge-LUBRICATION{ background: #1a3527; color: #4ade80; }
  .badge-PENUNJANG  { background: #3b2500; color: #fb923c; }
  .badge-AIRMATIC   { background: #2d1a5e; color: #c084fc; }
  .badge-ELECTRICAL { background: #3b1515; color: #f87171; }
  .conf-HIGH   { color: #4ade80; font-weight:700; }
  .conf-MEDIUM { color: #fbbf24; font-weight:700; }
  .conf-LOW    { color: #f87171; font-weight:700; }
  .bar-row { display:flex; align-items:center; gap:10px; margin-bottom:8px; }
  .bar-label { width: 110px; font-size:0.78rem; color:#94a3b8; flex-shrink:0; }
  .bar-bg { flex:1; background:#0f172a; border-radius:4px; height:14px; overflow:hidden; }
  .bar-fill { height:100%; border-radius:4px; transition: width 0.6s ease; }
  .bar-val { width:40px; text-align:right; font-size:0.78rem; color:#94a3b8; flex-shrink:0; }
  .info-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; margin: 16px 0; }
  .info-item label { font-size: 0.72rem; color: #64748b; text-transform: uppercase; letter-spacing: 0.05em; display: block; margin-bottom: 4px; }
  .info-item span { font-size: 0.9rem; color: #e2e8f0; }
  .reason-box { background: #0f172a; border-left: 3px solid #3b82f6; padding: 10px 14px; border-radius: 4px; font-size: 0.85rem; color: #94a3b8; margin-top: 12px; }
  table { width: 100%; border-collapse: collapse; font-size: 0.82rem; }
  th { background: #0f172a; padding: 10px 12px; text-align: left; color: #60a5fa; font-weight: 600; border-bottom: 1px solid #334155; }
  td { padding: 9px 12px; border-bottom: 1px solid #1e293b; color: #cbd5e1; }
  tr:hover td { background: rgba(59,130,246,0.05); }
  .search-row { display: flex; gap: 8px; margin-bottom: 16px; flex-wrap:wrap; }
  .search-row input, .search-row select { flex: 1; min-width:140px; background: #0f172a; border: 1px solid #334155; color: #e2e8f0; padding: 8px 12px; border-radius: 8px; font-size: 0.9rem; outline: none; }
  .search-row input:focus, .search-row select:focus { border-color: #3b82f6; }
  .loading { display: none; text-align: center; padding: 32px; color: #60a5fa; }
  .spinner { width: 32px; height: 32px; border: 3px solid #1e3a5f; border-top-color: #3b82f6; border-radius: 50%; animation: spin 0.8s linear infinite; margin: 0 auto 12px; }
  @keyframes spin { to { transform: rotate(360deg); } }
  .stats-row { display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 24px; }
  .stat-card { flex: 1; min-width: 100px; background: #1e293b; border: 1px solid #334155; border-radius: 10px; padding: 14px; text-align: center; }
  .stat-card .num { font-size: 1.5rem; font-weight: 700; color: #60a5fa; }
  .stat-card .lbl { font-size: 0.68rem; color: #64748b; margin-top: 4px; text-transform: uppercase; }
</style>
</head>
<body>
<div class="header">
  <h1>Spare Part Image Recognition</h1>
  <p>Kenali kategori spare part dari foto menggunakan Gemini AI</p>
</div>
<div class="container">

  <div class="stats-row">
    <div class="stat-card"><div class="num" id="statTotal">-</div><div class="lbl">Total Item</div></div>
    <div class="stat-card"><div class="num" id="statBearing">-</div><div class="lbl">Bearings</div></div>
    <div class="stat-card"><div class="num" id="statLubri">-</div><div class="lbl">Lubrication</div></div>
    <div class="stat-card"><div class="num" id="statPenunjang">-</div><div class="lbl">Penunjang</div></div>
    <div class="stat-card"><div class="num" id="statAirmatic">-</div><div class="lbl">Airmatic</div></div>
    <div class="stat-card"><div class="num" id="statElectrical">-</div><div class="lbl">Electrical</div></div>
  </div>

  <div class="card">
    <h2>Upload Foto Spare Part</h2>
    <div class="drop-zone" id="dropZone" onclick="document.getElementById('fileInput').click()">
      <div class="icon">📷</div>
      <p>Klik atau drag &amp; drop gambar di sini</p>
      <div class="hint">PNG, JPG, JPEG, BMP, WEBP — maks 16 MB</div>
      <input type="file" id="fileInput" accept="image/*">
    </div>
    <div class="preview-wrap" id="previewWrap">
      <img id="previewImg" src="" alt="preview">
    </div>
    <div style="text-align:center; margin-top:16px;">
      <button class="btn btn-primary" id="analyzeBtn" disabled onclick="analyze()">
        Kenali Spare Part
      </button>
    </div>
    <div class="loading" id="loading">
      <div class="spinner"></div>
      <p>Menganalisis gambar dengan Gemini AI...</p>
    </div>
  </div>

  <div class="card result-box" id="resultBox">
    <h2>Hasil Pengenalan</h2>
    <div id="itemNameDisplay" style="font-size:1.8rem; font-weight:700; color:#f1f5f9; margin-bottom:12px;">-</div>
    <div style="display:flex; align-items:center; gap:12px; margin-bottom:16px;">
      <span id="catBadge" class="badge">-</span>
      <span id="confText">-</span>
      <span id="scoreText" style="font-size:0.85rem; color:#64748b;"></span>
    </div>
    <div class="info-grid">
      <div class="info-item"><label>Deskripsi Kategori</label><span id="itemDesc">-</span></div>
      <div class="info-item"><label>Total Item dalam Kategori</label><span id="catTotal">-</span></div>
    </div>
    <div class="reason-box" id="reasonBox" style="display:none;"></div>

    <h2 style="margin-top:20px; margin-bottom:12px;">Skor per Kategori</h2>
    <div id="barsContainer"></div>

    <h2 style="margin-top:20px;">Item dalam Kategori</h2>
    <table>
      <thead><tr><th>Kode</th><th>Nama Barang</th><th>Min Stock</th><th>Max Stock</th><th>Satuan</th></tr></thead>
      <tbody id="matchBody"></tbody>
    </table>
  </div>

  <div class="card">
    <h2>Cari Spare Part</h2>
    <div class="search-row">
      <input type="text" id="searchQ" placeholder="Cari nama atau kode..." onkeydown="if(event.key==='Enter')doSearch()">
      <select id="searchCat">
        <option value="">Semua Kategori</option>
        <option value="BEARINGS">BEARINGS</option>
        <option value="LUBRICATION">LUBRICATION</option>
        <option value="PENUNJANG">PENUNJANG</option>
        <option value="AIRMATIC">AIRMATIC</option>
        <option value="ELECTRICAL">ELECTRICAL</option>
      </select>
      <button class="btn btn-primary" onclick="doSearch()">Cari</button>
    </div>
    <table>
      <thead><tr><th>Kategori</th><th>Kode</th><th>Nama Barang</th><th>Min</th><th>Max</th><th>Satuan</th></tr></thead>
      <tbody id="searchBody"></tbody>
    </table>
  </div>
</div>

<script>
const CAT_COLORS = {BEARINGS:'#3b82f6',LUBRICATION:'#4ade80',PENUNJANG:'#fb923c',AIRMATIC:'#c084fc',ELECTRICAL:'#f87171'};

fetch('/api/items').then(r=>r.json()).then(d=>{
  const t=d.total_per_category;
  document.getElementById('statTotal').textContent=d.grand_total;
  document.getElementById('statBearing').textContent=t['BEARINGS']||0;
  document.getElementById('statLubri').textContent=t['LUBRICATION']||0;
  document.getElementById('statPenunjang').textContent=t['PENUNJANG']||0;
  document.getElementById('statAirmatic').textContent=t['AIRMATIC']||0;
  document.getElementById('statElectrical').textContent=t['ELECTRICAL']||0;
});

const dz=document.getElementById('dropZone');
dz.addEventListener('dragover',e=>{e.preventDefault();dz.classList.add('drag-over');});
dz.addEventListener('dragleave',()=>dz.classList.remove('drag-over'));
dz.addEventListener('drop',e=>{e.preventDefault();dz.classList.remove('drag-over');handleFile(e.dataTransfer.files[0]);});
document.getElementById('fileInput').addEventListener('change',e=>handleFile(e.target.files[0]));

let selectedFile=null;
function handleFile(f){
  if(!f) return;
  selectedFile=f;
  const reader=new FileReader();
  reader.onload=e=>{
    document.getElementById('previewImg').src=e.target.result;
    document.getElementById('previewWrap').style.display='block';
    document.getElementById('analyzeBtn').disabled=false;
  };
  reader.readAsDataURL(f);
}

async function analyze(){
  if(!selectedFile) return;
  document.getElementById('loading').style.display='block';
  document.getElementById('resultBox').style.display='none';
  document.getElementById('analyzeBtn').disabled=true;

  const fd=new FormData();
  fd.append('image',selectedFile);

  try{
    const res=await fetch('/api/recognize',{method:'POST',body:fd});
    const data=await res.json();
    if(data.error){alert('Error: '+data.error);return;}

    const rec=data.recognition;
    const cat=data.category;
    document.getElementById('itemNameDisplay').textContent=rec.item_name||cat;
    const badge=document.getElementById('catBadge');
    badge.textContent=cat;
    badge.className='badge badge-'+cat;

    const conf=rec.confidence||'';
    document.getElementById('confText').innerHTML=`Confidence: <span class="conf-${conf}">${conf}</span>`;
    document.getElementById('scoreText').textContent=`(${rec.score}%)`;
    document.getElementById('itemDesc').textContent=rec.item_description||'-';
    document.getElementById('catTotal').textContent=data.total_items_in_category+' item';

    const reasonBox=document.getElementById('reasonBox');
    if(rec.reason){reasonBox.textContent='Alasan: '+rec.reason;reasonBox.style.display='block';}
    else{reasonBox.style.display='none';}

    const bars=document.getElementById('barsContainer');
    bars.innerHTML='';
    Object.entries(rec.scores_per_category||{}).sort((a,b)=>b[1]-a[1]).forEach(([c,v])=>{
      const color=CAT_COLORS[c]||'#64748b';
      bars.innerHTML+=`<div class="bar-row">
        <div class="bar-label">${c}</div>
        <div class="bar-bg"><div class="bar-fill" style="width:${Math.min(v,100)}%;background:${color}"></div></div>
        <div class="bar-val">${v}%</div>
      </div>`;
    });

    const tbody=document.getElementById('matchBody');
    tbody.innerHTML='';
    (data.matched_items||[]).forEach(item=>{
      const tr=document.createElement('tr');
      tr.innerHTML=`<td>${item.code||'-'}</td><td>${item.name||'-'}</td><td>${item.min_stock||'-'}</td><td>${item.max_stock||'-'}</td><td>${item.satuan||'-'}</td>`;
      tbody.appendChild(tr);
    });
    if(!data.matched_items||data.matched_items.length===0){
      tbody.innerHTML='<tr><td colspan="5" style="text-align:center;color:#64748b;">Tidak ada item</td></tr>';
    }
    document.getElementById('resultBox').style.display='block';
    document.getElementById('resultBox').scrollIntoView({behavior:'smooth'});
  }catch(e){
    alert('Error: '+e.message);
  }finally{
    document.getElementById('loading').style.display='none';
    document.getElementById('analyzeBtn').disabled=false;
  }
}

async function doSearch(){
  const q=document.getElementById('searchQ').value;
  const cat=document.getElementById('searchCat').value;
  const data=await fetch(`/api/search?q=${encodeURIComponent(q)}&category=${cat}`).then(r=>r.json());
  const tbody=document.getElementById('searchBody');
  tbody.innerHTML='';
  if(!data.results||data.results.length===0){
    tbody.innerHTML='<tr><td colspan="6" style="text-align:center;color:#64748b;">Tidak ada hasil</td></tr>';
    return;
  }
  data.results.forEach(item=>{
    const tr=document.createElement('tr');
    tr.innerHTML=`<td><span class="badge badge-${item.category}">${item.category}</span></td><td>${item.code||'-'}</td><td>${item.name||'-'}</td><td>${item.min_stock||'-'}</td><td>${item.max_stock||'-'}</td><td>${item.satuan||'-'}</td>`;
    tbody.appendChild(tr);
  });
}
</script>
</body>
</html>
"""

if __name__ == '__main__':
    print("=" * 55)
    print("  Spare Part Image Recognition — Gemini AI")
    print("=" * 55)
    for cat, items in SPARE_PARTS.items():
        print(f"  {cat}: {len(items)} items")
    print("=" * 55)
    print("  Buka browser: http://localhost:5000")
    print("=" * 55)
    app.run(debug=False, host='0.0.0.0', port=5000, use_reloader=False)
